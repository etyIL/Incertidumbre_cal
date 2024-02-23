# -*- coding: utf-8 -*-
"""
Created on Fri Dec  1 15:59:02 2023

@author: Estefania
"""

from tkinter import *
from tkinter import filedialog
import pandas as pd
import xlrd
import math
import openpyxl
import numpy as np
import xlsxwriter
import os
import os.path as path
from datetime import timedelta
from PIL import ImageTk, Image
import jinja2
import pathlib

os.chdir('.')

ventana= Tk()
ventana.title("Cálculo incertidumbre de balanzas")
ventana.resizable(1,1)
ventana.geometry("400x150")
ventana.iconbitmap('equilibrio-balanza-izquierda.ico')
ventana.config(bg="lightblue")

# cuadro de texto donde se verá la ruta del archivo a evaluar.
ruta= StringVar()

def calculo_incer(archivo):
    """ toma los datos de la planilla y realiza los cálculos pertinentes """

    info_balanza = archivo.iloc[:,[0,1,2,3,4]]
    fecha_cal = info_balanza.iloc[0,1]
    fecha_prox = fecha_cal + timedelta(days=365)
    fecha_prox = fecha_prox.strftime("%d/%m/%Y")
    responsable = info_balanza.iloc[1,1]
    n_cal = info_balanza.iloc[0,4]
    marca_bal = info_balanza.iloc[3,1]
    modelo_bal = info_balanza.iloc[4,1]
    n_serie = info_balanza.iloc[5,1]
    ubicacion = info_balanza.iloc[6,1]
    geo_plato = info_balanza.iloc[7,1]
    rango_max = info_balanza.iloc[8,1]
    resolucion = info_balanza.iloc[9,1]
    parte_decimal, parte_entera = math.modf(resolucion)

    def contar_d(n):
        """ cuenta los decimales de la resolución de la balanza """ 
        
        
        parte_decimal, parte_entera = math.modf(n)
        decimales = 0 
        if parte_decimal == 0:
            decimales = 0
        else:
            while (0<parte_decimal<1):
                decimales += 1
                parte_decimal = parte_decimal*10   
        return decimales

    decimales = contar_d(resolucion)

    def con_dec(n, decimales=decimales):
        """ aplica la cantidad de decimales necesarios (según los de la resolución)
        a los resultados numéricos de los cálculos"""
        
        n = f'{n:.{decimales}f}'
        return n

    # ubico datos
    
    T1 = info_balanza.iloc[19,1]
    T1_ = con_dec(T1,1)
    T2 = info_balanza.iloc[20,1]
    T2_= con_dec(T2,1)
    delta_T = abs(T1-T2)
    delta_T_= con_dec(delta_T,1)
    pesas_n = list(info_balanza.iloc[[4,5,6,7,8,9], 4]) # identifico el conjunto de pesas utilizadas en los ensayos
    pesas_n = [valor for valor in pesas_n if math.isnan(valor)==False] # elimino los espacios vacíos
    H1 = info_balanza.iloc[19,4]
    H1_= con_dec(H1,1)
    H2 = info_balanza.iloc[20,4]
    H2_= con_dec(H2,1)
    promedio_h = (H1+H2)/2
    promedio_h_ = con_dec(promedio_h,1)
    P1 = info_balanza.iloc[19,2]
    P1_ = con_dec(P1,1)
    P2 = info_balanza.iloc[20,2]
    P2_ = con_dec(P2,1)
    promedio_p = (P1+P2)/2
    promedio_p_= con_dec(promedio_p,1)
    
    # se ubican y extraen datos de repetibilidad y corrección...
    repetibilidad = pd.read_excel(archivo, engine= 'openpyxl', sheet_name= "Repetibilidad y corrección", header=[0], usecols="D:I", skiprows=(1,15))
    repetibilidad = repetibilidad.apply(lambda x: pd.Series(x.dropna().values))
    
    # identifico los masas utilizadas
    nominal_comb = repetibilidad.iloc[2,[0,1,2,3,4,5]] 
    pesas_comb = nominal_comb.values
    pesas_comb = list(pesas_comb)
    
    # elimino los espacios vacíos
    pesas_comb = [valor for valor in pesas_comb if math.isnan(valor)==False] 
    pesas_comb = sorted(pesas_comb)
    n = len(pesas_comb) 
    
    # lista de las pesas y sus combinaciones (si existieran)
    nominal = repetibilidad.iloc[1,:n] 
    nominal = list(nominal)
    
    #################################################
    # CÁLCULO DE REPETIBILIDAD
    
    correccion = repetibilidad.iloc[3,[0,1,2,3,4,5]]
    correccion_valores = correccion.values
    correccion_valores = list(correccion_valores)
    correccion_valores = [valor for valor in correccion_valores if math.isnan(valor)==False]# elimino los espacios vacíos

    # datos para el cálculo de repetibilidad...
    datos = repetibilidad.dropna(axis = "columns")
    datos_rep = datos.values
    
    
    # elimino el dato de la masa patrón y el encabezado del array porque solo necesito los datos de pesada tomados con esa pesa...
    datos_rep = np.delete(datos_rep, [0,1,2])  

    # calculo la desv. std. (repetibilidad) y el promedio de los datos tomados durante el ensayo con la masa ref. elegida...    
    prom = (sum(datos_rep))/len(datos_rep)
    prom = round(prom, decimales)
    
    # REPETIBILIDAD
    desviacion_std = datos_rep.std() 
    desviacion_std = round(desviacion_std, decimales)
    
    ###################################################
    # DATOS ASOCIADOS A LA MASA PATRÓN

    contribucion_masa_pat = pd.read_excel("Incertidumbre pesas.xlsx", engine= 'openpyxl', sheet_name= "Hoja2", header=[1])
    
    # incertidumbre por corrección de patrón...
    umc = []
    for pesa in pesas_comb:
        umc.append(contribucion_masa_pat.loc[contribucion_masa_pat['Masa ref.'] == pesa]['umc (error máx/raiz(3))'])    

    # incertidumbre por deriva...
    umd = []
    for pesa in pesas_comb: 
        umd.append(contribucion_masa_pat.loc[contribucion_masa_pat['Masa ref.'] == pesa]['umd (error máx/3*raiz(3))'])

    # incertidumbre por empuje del aire...
    umb = []
    for pesa in pesas_comb: 
        umb.append(contribucion_masa_pat.loc[contribucion_masa_pat['Masa ref.'] == pesa]['umb (0.1*(δ0/δc) + (error máx/4*√3))'])

    # incertidumbre por efectos convectivos...
    umconv = []
    for pesa in pesas_comb:
        if delta_T == 0: 
            umconv.append(contribucion_masa_pat.loc[contribucion_masa_pat['Masa ref.'] == pesa]['u (mconv) = ∆mconv/√3, deltaT=0'])
        if delta_T == 1:
            umconv.append(contribucion_masa_pat.loc[contribucion_masa_pat['Masa ref.'] == pesa]['u (mconv) = ∆mconv/√3, deltaT=0,5-1,4; 1'])
        if delta_T == 2:
            umconv.append(contribucion_masa_pat.loc[contribucion_masa_pat['Masa ref.'] == pesa]['u (mconv) = ∆mconv/√3, deltaT=1,5-2,4; 2']) 
        if delta_T == 3:
            umconv.append(contribucion_masa_pat.loc[contribucion_masa_pat['Masa ref.'] == pesa]['u (mconv) = ∆mconv/√3, deltaT=2,5-3,4; 3'])
    
    ####################################################  
    # CORRECCIÓN   
    
    # saco el valor medido al pesar la masa elegida para repetibilidad porque no se utiliza para corrección...
    for valor in correccion_valores: 
        if valor in datos_rep:
            medida_a_eliminar = correccion_valores.index(valor)
            correccion_ = correccion_valores.pop(medida_a_eliminar)
            
    # agrego en la lista el valor promedio de las mediciones de la pesa elegida para repetibilidad.
    correccion_valores.append(prom)
    correccion_valores = sorted(correccion_valores) #ordeno de menor a mayor

    # coloco los decimales correspondientes haciendo uso de la función
    correccion__valores = []
    for c in correccion_valores:
        c = round(c, decimales)
        c = con_dec(c, decimales)
        correccion__valores.append(c)
        
    # leo los datos de los errores y correcciones asociados a las pesas ref...
    incer_pesas = pd.read_excel("Incertidumbre pesas.xlsx", engine= "openpyxl", sheet_name = "Hoja1", header = [0], usecols= "A,C,D,E", skiprows = (1))

    # ubico los datos de corrección de las pesas empleadas en los ensayos y los guardo en una lista...
    patron_corregido = []
    for pesa in pesas_comb:
        patron_corregido.append(incer_pesas.loc[incer_pesas['Valor Nominal [g]'] == pesa]['Patrón corregido'])
    
    # número de certificado de las pesas (dato no muy relevante)
    n_certificado = []   
    for pesa in pesas_n:
        n_certificado.append(incer_pesas.loc[incer_pesas['Valor Nominal [g]'] == pesa]['N Certificado'])

    # guardo en una lista la diferencia entre el dato de corrección tomado en el ensayo y la masa patrón corregida...
    dif = []
    for v1, v2 in zip(correccion_valores, patron_corregido):
        dif.append(v1-v2)
    
    # paso la lista a un array para poder manipularlo
    dif_arr = np.array(dif)
    
    # extraigo el máximo valor del array...
    error_max = (max(abs(dif_arr))) 
    max_err = error_max.item()
    
    # le coloco los decimales correspondientes...
    error__max = round(max_err, decimales)
    error__max = con_dec(max_err, decimales)

    #####################################################
    #EXCENTRICIDAD
    
    # se ubican y extraen los datos del ensayo de excentricidad...
    datos_excen = pd.read_excel(archivo, engine= 'openpyxl', sheet_name= "Ensayo Excentricidad", header=[0], usecols="C:D")
    valores_excen = datos_excen.iloc[12:19]
    valores_excen = valores_excen.apply(lambda x: pd.Series(x.dropna().values))
    valor_nominal = valores_excen.iloc[0,0]
    lectura_excen = valores_excen.iloc[1:, 1]
    dif_excen = lectura_excen.values

    dif__excen = []
    for d in dif_excen:
        d = round(d, decimales)
        d = con_dec(d, decimales)
        dif__excen.append(d)
    
    # se extrae el máximo de los datos y se calcula la excentricidad...
    max_excen = max(abs(dif_excen))
    excen = (max_excen/(2*correccion_*(3**(1/2))))
    
    # calculo la excentricidad de cada pesa utilizada y lo guardo en una lista...
    excen_pesas = []
    for pesa in pesas_comb:
        excen_pesas.append(excen*pesa)
        
    #############################################
    # ÚLTIMOS CÁLCULOS
    
    # calculo la división de escala   
    division_escala = (resolucion)/(2*(3**(1/2)))

    # calculo, por un lado, para cada pesa, las contribuciones de la excentricidad, umc, umb, umd, umconv...
    punto_cal = []  
    for z1, z2, z3, z4, z5 in zip(excen_pesas, umc, umb, umd, umconv):
        punto_cal.append((z1**2) + (z2**2) + (z3**2) + (z4**2) + (z5**2))
        
    # por el otro, a cada valor de la lista anterior, además le sumo la contribución de la repetibilidad y dos veces la división de escala
    dato_cal = []
    for p in punto_cal:
        p = (p + (desviacion_std**2) + 2*(division_escala**2))**(1/2)
        dato_cal.append(p)
        
    # paso la lista a array para poder manipular los datos   
    dato_cal_arr = np.array(dato_cal) 
    
    # el máximo de la lista ya es el valor de incertidumbre de la balanza!!
    dato_incer = max(dato_cal_arr*2) 
    incer_d = dato_incer.item()
    
    # calculo la tolerancia
    tolerancia = 3*(incer_d)

    def necesaria_correccion(tolerancia, max_err):
        """ evalúa si se necesita corrección de escala de la balanza
        en función de los valores de error máximo y tolerancia.
        Devuelve SI o NO"""
        
        if tolerancia < max_err:
           necesaria_correccion = 'SI'
        else:
           necesaria_correccion = 'NO'
        return necesaria_correccion
        
    necesaria_correccion = necesaria_correccion(tolerancia, error_max)   
    
    # le coloco los decimales correspondientes...
    tolerancia_ = round(tolerancia, decimales)
    tolerancia_ = con_dec(tolerancia, decimales)

    # ECUACIONES DE LOS RANGOS
    
    # ecuacion primer rango...
    pend_rango1 = (dif_arr[0]-0)/(pesas_comb[0]-0)
    ord_rango1 = dif_arr[0]-(pend_rango1*pesas_comb[0])
    pend_rango11 = pend_rango1.item()
    pend_rango11 = round(pend_rango11, 4)
    ord_rango11 = ord_rango1.item()
    ord_rango11 = round(ord_rango11, 4)

    # ciclo para cálculo de pendientes de los rangos
    pend_rango_= []
    i = 0
    while i+1 < len(dif_arr):
        pend_rango_.append((dif_arr[i+1]-dif_arr[i])/(pesas_comb[i+1]-pesas_comb[i]))
        i += 1
        
    pend__rango_= []
    for p in pend_rango_:
        pi = p.item()
        pi = round(pi, 6)
        #p = con_dec(p, decimales)
        pend__rango_.append(pi)

    # ciclo para cálculo de ordenadas de los rangos
    ord_rango_ = []
    j = 0
    while j+1 < len(dif_arr):
        ord_rango_.append(dif_arr[j+1]-(pend_rango_[j]*pesas_comb[j+1]))
        j+=1
          
    ord__rango_ = []
    for o in ord_rango_:
        oi = o.item()
        oi = round(oi, 6)
        ord__rango_.append(oi)
        
    ##################################################
    # EXPORTO LOS RESULTADOS AL CERTIFICADO EN FORMATO EXCEL YA EXISTENTE
    
    certificado_excel = openpyxl.load_workbook('Certificado.xlsx')
    certificado = certificado_excel['Certificado de Calibración']
    certificado['J4'] = fecha_cal
    certificado['J5'] = fecha_prox
    certificado['C6'] = responsable
    certificado['C7'] = ubicacion
    certificado ['C8'] = 'nombre institución'
    certificado['B11'] = marca_bal
    certificado['H11'] = modelo_bal
    certificado['B12'] = rango_max
    resolucion_= con_dec(resolucion,decimales)
    certificado['H12'] = resolucion_
    certificado['C15'] = 'SGC.CAL.PR.01.v01'
    certificado['C16'] = 'gramo'

    for pesa in pesas_n:
        if pesa<1:
            posicion = pesas_n.index(pesa)
            pesa_ = con_dec(pesa,3)
            pesas_n[posicion] = pesa_
        
    m = len(pesas_n)-1 
        
    for rows in certificado.iter_cols(min_col=6, max_col=6, min_row=21, max_row=(21+m)):
      for r, row in enumerate(rows):
          row.value = pesas_n[r]

    n_certificado = np.array(n_certificado)
    n__certificado = []
    for i in n_certificado:
        i = f'{i}'
        n__certificado.append(i)
        
    for rows in certificado.iter_cols(min_col=7, max_col=7, min_row=21, max_row=(21+m)):
      for r, row in enumerate(rows):
          row.value = n__certificado[r]
           
    certificado['F32'] = T1_
    certificado['G32'] = T2_
    certificado['F33'] = H1_
    certificado['G33'] = H2_
    certificado['F34'] = P1_
    certificado['G34'] = P2_

    datos__rep = []
    for d in datos_rep:
        d = round(d, decimales)
        d = con_dec(d, decimales)
        datos__rep.append(d)

    for rows in certificado.iter_cols(min_col=3, max_col=3, min_row=38, max_row=47):
      for r, row in enumerate(rows):
          row.value = datos__rep[r]

    prom_ = con_dec(prom,decimales)
    desviacion_std_ = con_dec(desviacion_std)
    certificado['I38'] = prom_
    certificado['I40'] = desviacion_std_

    for rows in certificado.iter_cols(min_col=2, max_col=2, min_row=54, max_row=58):
      for r, row in enumerate(rows):
          row.value = dif__excen[r]
          
    certificado['A52'] = valor_nominal

    if geo_plato == 'circular':
        certificado['H51'] = 'CIRCULAR         X'
    if geo_plato == 'triangular':
        certificado['G51'] = 'TRIANGULAR  X'
    if geo_plato == 'rectangular':
        certificado['I51'] = 'RECTANGULAR X'

    for n in nominal:
        if type(n) == float:
            lugar_n = nominal.index(n)
            n_ = con_dec(n,3)
            nominal[lugar_n] = n_

    n = len(nominal)-1          
    for rows in certificado.iter_cols(min_col=1, max_col=1, min_row=63, max_row=(63+n)):
      for r, row in enumerate(rows):
          row.value = nominal[r]          

    patron_corregido = np.array(patron_corregido)
    patron__corregido = []
    for i in patron_corregido:
        ii = i.item()
        ii = round(ii,decimales)
        ii = con_dec(ii, decimales)
        patron__corregido.append(ii)
        
    for rows in certificado.iter_cols(min_col=2, max_col=2, min_row=63, max_row=(63+n)):
       for r, row in enumerate(rows):
           row.value = patron__corregido[r] 

    for rows in certificado.iter_cols(min_col=5, max_col=5, min_row=63, max_row=(63+n)):
       for r, row in enumerate(rows):
           row.value = correccion__valores[r]

    dif__arr = []
    for d in dif_arr:
        di = d.item()
        di = con_dec(di, decimales)
        dif__arr.append(di)

    for rows in certificado.iter_cols(min_col=7, max_col=7, min_row=63, max_row=(63+n)):
       for r, row in enumerate(rows):
           row.value = dif__arr[r]

    dato__cal__arr = []       
    for dato in dato_cal_arr:
        datoi = dato.item()
        datoi = 2*datoi
        datoi = round(datoi, decimales)
        datoi = con_dec(datoi, decimales)
        dato__cal__arr.append(datoi)

    for rows in certificado.iter_cols(min_col=9, max_col=9, min_row=63, max_row=(63+n)):
       for r, row in enumerate(rows):
           row.value = dato__cal__arr[r]
          
    dato__incer = round(incer_d, decimales)
    dato__incer = con_dec(dato__incer, decimales)
    
    f_dif = []
    for d in dif_arr:
        di = d.item()
        f_dif.append(di)
        
    
    if (abs(f_dif[0]))>tolerancia:
        certificado['B80'] = f'y = {pend_rango1}*X + {ord_rango1}'
        certificado['A80'] = f' 0 - {pesas_comb[0]}'
    else:
        pass

    z = len(f_dif)-2

    for rows in certificado.iter_cols(min_col=2, max_col=2, min_row=81, max_row=(81+z)):
        for r, row in enumerate(rows):
            if (abs(f_dif[r+1]))>tolerancia:
                row.value = f'y = {pend__rango_[r]}*X + {ord__rango_[r]}' 
            else:
                pass

    for rows in certificado.iter_cols(min_col= 1, max_col= 1, min_row=81, max_row=(81+z)):
        for r, row in enumerate(rows):
            if (abs(f_dif[r+1]))>tolerancia:
                row.value = f'{pesas_comb[r]} - {pesas_comb[r+1]}' 
            else:
                pass
        
    if necesaria_correccion == 'SI':
        certificado['A77'] = 'APLICA  X'
    else:
        certificado['B77'] = 'NO APLICA  X'
        
    certificado['I70'] = dato__incer
    certificado['I71'] = tolerancia_
    certificado['I72'] = error__max
    certificado['I73'] = necesaria_correccion
    
    # guardo el ceryificado con el nombre de la balanza correspondiente
    Certificado= certificado_excel.save('Certificado ' + f'{marca_bal} ' + f'{ubicacion} ' + 'rango ' + f'{rango_max}' + '.xlsx')
    os.startfile('Certificado '+ f'{marca_bal} ' + f'{ubicacion} ' + 'rango ' + f'{rango_max}' + '.xlsx' )

    return Certificado
    ###############################################
    
def abrirarchivo():
    """ busca en el directorio el archivo a partir del cual procesar los datos.
    Si existe, lo lee y evalúa que sea el correspondiente y/o que esté completo, sino arroja error
    Si está OK, llama a la función para hacer los cálculos.
    Devuelve el archivo que se pasó por interfaz """
    
    directorio= pathlib.Path('.')
    direccion= filedialog.askopenfilename(title= "Abrir", initialdir= "C:/", filetypes= (("Archivos xlsx", "*.xlsx"), ("Todos los archivos", "*.*")))
    for file in directorio.iterdir():
        if path.exists(direccion):
    
            datos = pd.read_excel(direccion, engine= 'openpyxl', sheet_name= "Datos de la balanza", header=None, index_col= 0, skiprows=(0,3))
            archivo= pd.DataFrame(datos)
            consulta= archivo.iloc[9,1]
            if consulta!= float(consulta):
                ventana2= Tk()
                ventana2.title('Resultado')
                ventana2.resizable(1,1)
                ventana2.geometry("500x100")
                #ventana.iconbitmap("ruta foto")
                ventana2.config(bg="lightblue")
                texto= Label(ventana2, text="Hay un error en el procesamiento de los datos. Verifique que el archivo esté correcto", bg= 'lightblue')
                texto.pack()
                

            else:
                ruta.set(direccion)
                calculo_incer(archivo)
                ventana2= Tk()
                ventana2.title('Resultado')
                ventana2.resizable(1,1)
                ventana2.geometry("400x100")
                ventana2.iconbitmap('equilibrio-balanza-izquierda.ico')
                ventana2.config(bg="lightblue")
                texto= Label(ventana2, text="El Certificado ya está listo!", bg= 'lightblue')
                texto.pack()
    
        boton_cerrar = Button(ventana2, text= 'Cerrar', command= ventana2.destroy)
        boton_cerrar.pack()
        return archivo 
        
abrir= Button(ventana, text= "Abrir Archivo", command= abrirarchivo)
abrir.pack(pady= 10)

cuadroarchivo= Entry(ventana, textvariable=ruta, width=40)
cuadroarchivo.pack()  

def cerrar():
    ventana.destroy()
                                
cerrar= Button(ventana, text= 'Cerrar', command= cerrar)
cerrar.pack(pady=10)
ventana.mainloop()
 



